import logging
import os
import re
from datetime import datetime as dt
from typing import List, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from config import EXCEL_FOLDER
from data_structures import Dimension


def get_all_rows(sheet: Worksheet, dim: Dimension) -> List[Tuple]:
    return [vals for vals in sheet.iter_rows(1, dim.height, 1, dim.width, values_only=True)]


def get_latest_full_cell(row: tuple) -> int:
    for i, val in enumerate(reversed(row)):
        if not val:
            continue
        return len(row) - i


def get_width(rows: List[Tuple]) -> int:
    last_positions = [get_latest_full_cell(row) for row in rows if get_latest_full_cell(row)]
    return max(set(last_positions), key=last_positions.count)


def get_height(sheet: Worksheet, rows: List[Tuple]) -> int:
    return next((len(rows) - i for i, row in enumerate(reversed(rows)) if 'ВСЕГО' in row), sheet.max_row)


def get_dimensions(sheet: Worksheet, rows: List[Tuple]) -> Dimension:
    return Dimension(width=get_width(rows=rows), height=get_height(sheet, rows=rows))


def get_date(rows: List[Tuple]) -> str:
    months_dict = {1: 'JAN', 2: 'FEV', 3: 'MAR', 4: 'APR', 5: 'MAI', 6: 'IYN',
                   7: 'JYL', 8: 'AVG', 9: 'SEN', 10: 'OKT', 11: 'NOJ', 12: 'DEK'}
    date = dt.strptime(re.search(r'\d\d\.\d\d\.\d{4,}', rows[0][0]).group(), '%d.%m.%Y')
    return f'{months_dict[date.month]}_{date.year}'


def xl_copy(sheet_from: Worksheet, sheet_to: Worksheet, dimension: Dimension) -> None:
    for i in range(1, dimension.height + 1):
        for j in range(1, dimension.width + 1):
            cell = sheet_from.cell(row=i, column=j)
            _ = sheet_to.cell(row=i, column=j, value=cell.value)


def find_empty_columns(rows: List[Tuple]) -> List[int]:
    st = next(i + 1 for i, col in enumerate(rows[3]) if col)
    cols = rows[3][st::]
    return [i + 1 for i, col in enumerate(cols, start=st) if not col]


def remove_empty_cols(sheet: Worksheet, rows: List[Tuple]) -> None:
    empty_cols = find_empty_columns(rows=rows)
    for i, col_idx in enumerate(empty_cols):
        sheet.delete_cols(col_idx - i, 1)


def correct(excel_name: str) -> Tuple[str, str]:
    logging.info('Started correcting Excel file.')

    full_path = os.path.join(EXCEL_FOLDER, excel_name)

    workbook = openpyxl.load_workbook(filename=full_path, data_only=True)
    logging.info(f'Loaded workbook: {full_path}')
    default_sheet = workbook.worksheets[0]
    logging.info(f'Loaded default sheet: {default_sheet.title}')
    raw_dimension = Dimension(width=default_sheet.max_column, height=default_sheet.max_row)
    logging.info(f'Raw dimension: {raw_dimension}')
    rows = get_all_rows(default_sheet, raw_dimension)
    logging.info(f'Got all rows from default sheet. Rows: {len(rows)}')

    new_dimension = get_dimensions(sheet=default_sheet, rows=rows)
    logging.info(f'New dimension: {new_dimension}')
    rows = [row[0:new_dimension.width] for row in rows[:new_dimension.height]]
    logging.info(f'Got all new rows. New rows: {len(rows)}')
    new_workbook = openpyxl.Workbook()
    logging.info('Created new workbook.')
    new_sheet = new_workbook.worksheets[0]
    logging.info('Created new sheet.')
    date = get_date(rows=rows)
    logging.info(f'Got date: {date}')

    xl_copy(sheet_from=default_sheet, sheet_to=new_sheet, dimension=new_dimension)
    logging.info('Copied data from default sheet to new sheet.')
    remove_empty_cols(sheet=new_sheet, rows=rows)
    logging.info('Removed empty columns from new sheet.')
    corrected_excel_name = f'Prov_{date}.xlsx'
    corrected_excel_full_path = os.path.join(EXCEL_FOLDER, corrected_excel_name)
    logging.info(f'Corrected Excel name: {corrected_excel_name}')
    new_workbook.save(filename=corrected_excel_full_path)
    logging.info(f'Saved new workbook. Path: {corrected_excel_full_path}')
    logging.info('Finished correcting Excel file.')

    return corrected_excel_name, date
