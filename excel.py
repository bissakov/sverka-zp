import os
import re
import shutil
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime as dt
from typing import List
from data_structures import ExcelInfo, Dimension


class Excel:
    def __init__(self, excel: ExcelInfo) -> None:
        self.excel_path = excel.path
        self.excel_name = excel.name
        shutil.copyfile(os.path.join(self.excel_path, self.excel_name), os.path.join(r'C:', 'test.xlsx'))

        self.workbook = openpyxl.load_workbook(filename=os.path.join(r'C:', 'test.xlsx'), data_only=True)
        self.default_sheet = self.workbook.worksheets[0]
        raw_dimension = Dimension(width=self.default_sheet.max_column, height=self.default_sheet.max_row)
        self.rows = self.get_all_rows(self.default_sheet, raw_dimension)

        self.new_dimension = Dimension(width=self.get_width(), height=self.get_height())
        self.rows = [row[0:self.new_dimension.width] for row in self.rows[:self.new_dimension.height]]

        self.new_workbook = openpyxl.Workbook()
        self.new_sheet = self.new_workbook.worksheets[0]
        self.corrected_name = ''
        self.date = self.get_date()

    def get_date(self) -> str:
        months_dict = {1: 'JAN', 2: 'FEV', 3: 'MAR', 4: 'APR', 5: 'MAI', 6: 'IYN',
                       7: 'JYL', 8: 'AVG', 9: 'SEN', 10: 'OKT', 11: 'NOJ', 12: 'DEK'}
        date = dt.strptime(re.search(r'\d\d\.\d\d\.\d{4,}', self.rows[0][0]).group(), '%d.%m.%Y')
        return f'{months_dict[date.month]}_{date.year}'

    def correct(self) -> None:
        self.xl_copy()
        self.remove_empty_cols()
        self.corrected_name = f'Prov_{self.date}.xlsx'
        new_path = os.path.join(r'C:\Temp', self.corrected_name)
        open(new_path, 'w').close()
        self.new_workbook.save(filename=new_path)
        # shutil.copyfile(src=r'C:\Users\robot.ad\Desktop\Prov_JYL_2022.xlsx', dst=new_path)

    @staticmethod
    def get_all_rows(sheet: Worksheet, dim: Dimension) -> List[tuple]:
        return [vals for vals in sheet.iter_rows(1, dim.height, 1, dim.width, values_only=True)]

    @staticmethod
    def get_latest_full_cell(row: tuple) -> int:
        for i, val in enumerate(reversed(row)):
            if not val:
                continue
            return len(row) - i

    def get_width(self) -> int:
        last_positions = [self.get_latest_full_cell(row) for row in self.rows if self.get_latest_full_cell(row)]
        return max(set(last_positions), key=last_positions.count)

    def get_height(self) -> int:
        return next((len(self.rows) - i for i, row in enumerate(reversed(self.rows)) if 'ВСЕГО' in row), self.default_sheet.max_row)

    def xl_copy(self) -> None:
        for i in range(1, self.new_dimension.height + 1):
            for j in range(1, self.new_dimension.width + 1):
                cell = self.default_sheet.cell(row=i, column=j)
                _ = self.new_sheet.cell(row=i, column=j, value=cell.value)

    def find_empty_columns(self) -> List[int]:
        st = next(i + 1 for i, col in enumerate(self.rows[3]) if col)
        cols = self.rows[3][st::]
        return [i + 1 for i, col in enumerate(cols, start=st) if not col]

    def remove_empty_cols(self) -> None:
        empty_cols = self.find_empty_columns()
        for i, col_idx in enumerate(empty_cols):
            self.new_sheet.delete_cols(col_idx - i, 1)
